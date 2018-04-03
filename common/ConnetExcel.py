# -*- coding: utf-8 -*-

import os
import xlrd
import openpyxl

#获取path路径
compath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
exlpath = os.path.join(compath,"testData")

#参数数据datain  参考项目demo，时间关系，后续直接获取列数
idcol = 0
intcol = 1
bookidcol = 4
booknamecol = 5
bookauthorcol = 6
bookyearcol = 7
bookdigestcol = 8

#dataout
comResult = 3
booknumcol = 4
rescodecol = 5
resjsoncol = 6
bookidcol1 = 7
booknamecol1 = 8
bookauthorcol1 = 9
bookyearcol1 = 10
bookdigestcol1 = 11

class ConnectExcelData(object):
     def __init__(self, excelFile, excelPath):
        self.excelFile = excelFile
        self.excelPath = excelPath
        self.absFile = os.path.join(excelPath,excelFile)
        self.caseList = []

     def get_case_List(self, absFile, begincase, casenum):
        openExcel = xlrd.open_workbook(absFile)
        beginRow = 0
        try:
            table = openExcel.sheet_by_name('datain')
            trows = table.nrows
            for n in range(1, trows):
                if table.cell(n, idcol).value == begincase:
                    beginRow = n
                    print beginRow
                continue
            for n in range(beginRow,beginRow+casenum):
                tmdict = {}
                tmdict['id'] = table.cell(n, idcol).value
                tmdict['interface'] = table.cell(n, intcol).value
                tmdict['bookid'] = table.cell(n, bookidcol).value
                tmdict['bookname'] = table.cell(n, booknamecol).value
                tmdict['bookauthor'] = table.cell(n, bookauthorcol).value
                tmdict['bookyear'] = table.cell(n, bookyearcol).value
                tmdict['bookdigest'] = table.cell(n, bookdigestcol).value
                self.caseList.append(tmdict)
        except Exception, e:
            raise
        finally:
            pass
        return self.caseList

#获取case对应的行数
     def get_row_num(self, absFile, sheetName, case):
        openExcel = xlrd.open_workbook(absFile)
        therow = 0
        try:
            table = openExcel.sheet_by_name(sheetName)
            trows = table.nrows
            for n in range(1, trows):
                if table.cell(n, idcol).value == case:
                    therow = n
                continue
        except Exception, e:
            raise
        finally:
            pass
        return therow

#更新表格数据

     def write_case_result(self, resjson, restime, booknum, rescode, absFile, therow):
        writeExcel = openpyxl.load_workbook(absFile)
        try:
            wtable = writeExcel.get_sheet_by_name('dataout')
            wtable.cell(row=therow+1, column=comResult+1, value=restime)
            wtable.cell(row=therow+1, column=booknumcol+1, value=booknum)
            wtable.cell(row=therow + 1, column=rescodecol + 1, value=rescode)
            wtable.cell(row=therow + 1, column=resjsoncol + 1, value=resjson)
            writeExcel.save(absFile)
        except Exception, e:
            raise
        finally:
            pass

     def write_book_result(self, book, absFile, therow):
        writeExcel = openpyxl.load_workbook(absFile)
        try:
            wtable = writeExcel.get_sheet_by_name('dataout')
            wtable.cell(row=therow+1, column=bookidcol1+1, value=book['id'])
            wtable.cell(row=therow+1, column=booknamecol1+1, value=book['name'])
            wtable.cell(row=therow+1, column=bookauthorcol1+1, value=book['author'])
            wtable.cell(row=therow+1, column=bookyearcol1 + 1, value=book['year'])
            wtable.cell(row=therow+1, column=bookdigestcol1 + 1, value=book['digest'])
            writeExcel.save(absFile)
        except Exception, e:
            raise
        finally:
            pass

"""
if __name__ == '__main__':
    filename = 'caseData.xlsx'
#    absFile = os.path.join(filename, exlpath)
    begincase = '00001'
    casenum=1
    conxls = ConnectExcelData(filename, exlpath)
    caselist = conxls.get_case_List(conxls.absFile, begincase, casenum)
    for casedict in caselist:
        caseid = casedict['id']
        print caseid
        caseintf = casedict['interface']
        print caseintf

        resjson = 0
        restime = '0001'
        booknum = 5
        therow = conxls.get_row_num(conxls.absFile, 'dataout', caseid)
        print therow
        conxls.write_case_result(resjson, restime, booknum, conxls.absFile, therow+1)
        print therow
"""


